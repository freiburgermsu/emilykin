#!/usr/bin/env python3
"""
Populate Table S2a (RPKM per MAG), S2b (RPKM per sample), S3 (gene copy number)
in MAG_target_gene_empty_tables.xlsx. K00376 is split into nosZ I / nosZ II using
the tree-based clade calls (clade_classify/out/allbins_nosz_master.tsv). Rows are
matched by MAG name in column A; the prefilled iterative-ID (col B) and S3 taxonomy
formula (col C) are left untouched. Saves a populated COPY (template preserved).
"""
import csv
from collections import defaultdict
import openpyxl

ROOT = "/home/freiburger/Documents/EmilyKin/"
SRC_XLSX = ROOT + "MAG_target_gene_empty_tables.xlsx"
OUT_XLSX = SRC_XLSX   # populate in place (same workbook)
SAMPLES = ["CAN_1","CAN_2","CAN_3","CAN_4","CAN_5"]

def read_tsv(p, delim="\t"):
    with open(p) as f:
        return list(csv.DictReader(f, delimiter=delim))

# ── KO order (authoritative) from the copy-number matrix header ──────────────
cn_rows = read_tsv(ROOT+"ko_copy_number_matrix.csv", ",")
KOS = [k for k in cn_rows[0].keys() if k != "MAG"]          # 42 KOs incl K00376
rpkm_comb = read_tsv(ROOT+"ko_gene_abundance/ko_rpkm_combined.tsv")
assert [k for k in rpkm_comb[0] if k.startswith("K")] == KOS, "KO order mismatch!"
# build the output column order: K00376 -> (nosZ I, nosZ II)
COLSEQ = []
for k in KOS:
    COLSEQ += (["nosZ I","nosZ II"] if k == "K00376" else [k])

# ── per-MAG copy number + RPKM(combined) ─────────────────────────────────────
copy = {r["MAG"]: r for r in cn_rows}
rpkmM = {r["MAG"]: r for r in rpkm_comb}

# ── nosZ clade split per MAG (tree calls) ────────────────────────────────────
nosz = {}   # MAG -> (cladeI, cladeII)
for r in read_tsv(ROOT+"clade_classify/out/allbins_nosz_master.tsv"):
    nosz[r["MAG"]] = (int(r["cladeI"]), int(r["cladeII"]))
def split_frac(mag):
    cI, cII = nosz.get(mag, (0,0))
    tot = cI + cII
    if tot == 0: return None          # no clade info
    return cI/tot, cII/tot

# ── per-(ko,mag,sample) RPKM for S2b community totals ────────────────────────
persamp = defaultdict(lambda: defaultdict(float))   # (ko) -> sample -> sum over MAGs
k376_persamp = defaultdict(lambda: defaultdict(float))  # mag -> sample -> K00376 rpkm
for r in read_tsv(ROOT+"ko_gene_abundance/ko_rpkm_per_sample.tsv"):
    ko, mag = r["ko"], r["mag"]
    for s in SAMPLES:
        v = float(r[s])
        if ko == "K00376": k376_persamp[mag][s] += v
        else:              persamp[ko][s] += v

# ── reporting: edge cases ────────────────────────────────────────────────────
k376_rpkm_mags = [m for m in rpkmM if float(rpkmM[m].get("K00376",0) or 0) > 0]
unsplit = [m for m in k376_rpkm_mags if split_frac(m) is None]
print(f"MAGs: {len(copy)} | nosZ-classified (allbins): {len(nosz)} | "
      f"K00376 RPKM>0: {len(k376_rpkm_mags)} | of those without clade split: {len(unsplit)} {unsplit}")
cn_mismatch = [m for m in copy if (int(copy[m]['K00376']) != sum(nosz.get(m,(0,0))))]
print(f"MAGs where matrix K00376 copies != allbins(cladeI+cladeII): {len(cn_mismatch)} (using allbins for nosZ I/II)")

def nosz_split_value(mag, total, kind):
    """Split a total (RPKM) into clade I/II; if unclassified, assign all to nosZ II."""
    fr = split_frac(mag)
    if fr is None: return (0.0 if kind=="I" else total)
    return total * (fr[0] if kind=="I" else fr[1])

# ── populate workbook ────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(SRC_XLSX)
def num(x):
    try: return round(float(x), 4)
    except (TypeError, ValueError): return 0.0

# S2a — RPKM per MAG (data cols start at C=3)
ws = wb["Table S2a_RPKM per MAG"]; filled=0; miss=[]
for row in range(4, ws.max_row+1):
    mag = ws.cell(row,1).value
    if not mag: continue
    if mag not in rpkmM: miss.append(mag); continue
    r = rpkmM[mag]; k376 = num(r.get("K00376"))
    for j,col in enumerate(COLSEQ):
        c = 3+j
        if col == "nosZ I":   ws.cell(row,c, round(nosz_split_value(mag,k376,"I"),4))
        elif col == "nosZ II":ws.cell(row,c, round(nosz_split_value(mag,k376,"II"),4))
        else:                 ws.cell(row,c, num(r.get(col)))
    filled+=1
print(f"S2a filled {filled} MAG rows; unmatched={miss}")

# S3 — copy number per MAG (data cols start at D=4; C is taxonomy formula, leave it)
ws = wb["Table S3_gene copies"]; filled=0; miss=[]
for row in range(4, ws.max_row+1):
    mag = ws.cell(row,1).value
    if not mag: continue
    if mag not in copy: miss.append(mag); continue
    r = copy[mag]; cI,cII = nosz.get(mag,(0,0))
    for j,col in enumerate(COLSEQ):
        c = 4+j
        if col == "nosZ I":   ws.cell(row,c, cI)
        elif col == "nosZ II":ws.cell(row,c, cII)
        else:                 ws.cell(row,c, int(r.get(col,0) or 0))
    filled+=1
print(f"S3 filled {filled} MAG rows; unmatched={miss}")

# S2b — RPKM per sample (community totals; data cols start at B=2)
ws = wb["Table S2b_RPKM per sample"]
for row in range(4, ws.max_row+1):
    s = ws.cell(row,1).value
    if s not in SAMPLES: continue
    # nosZ split: sum over MAGs of clade-fractioned K00376 per-sample RPKM
    nI = sum(nosz_split_value(m, k376_persamp[m][s], "I") for m in k376_persamp)
    nII= sum(nosz_split_value(m, k376_persamp[m][s], "II") for m in k376_persamp)
    for j,col in enumerate(COLSEQ):
        c = 2+j
        if col == "nosZ I":   ws.cell(row,c, round(nI,4))
        elif col == "nosZ II":ws.cell(row,c, round(nII,4))
        else:                 ws.cell(row,c, round(persamp[col][s],4))
print(f"S2b filled {len([1 for r in range(4,ws.max_row+1) if ws.cell(r,1).value in SAMPLES])} sample rows")

wb.save(OUT_XLSX)
print(f"\nsaved -> {OUT_XLSX}")
