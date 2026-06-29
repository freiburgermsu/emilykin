#!/usr/bin/env python3
"""
Fill the "Table S6_ANI" worksheet of MAG_target_gene_filled.xlsx (ANI novelty table).

Two-stage design:
  * LOCAL stage (always runs): fills the columns derivable from data already in
    the repo — MAG iteration ID, full GTDB taxonomy string, the six rank columns
    (Phylum..Species), Reference database, and same-GTDB-species co-member notes.
  * ANI stage (runs iff the GTDB-Tk classify summary is present): fills closest
    reference name/accession, ANI, AF, the species-level call, the novelty
    interpretation, and appends ANI provenance to Notes.

The GTDB-Tk summary is NOT in the repo; it lives on the compute cluster at
    /scratch1/afreiburger/emilykin/processed/mag/gtdbtk/classify/gtdbtk.bac120.summary.tsv
Drop that file into the repo root (or pass --gtdbtk) and re-run to complete S6.

Novelty rules (collaborator-specified; AF is a fraction 0-1, ANI a percent):
    ANI >= 95 and AF >= 0.60  -> same species-level cluster as closest reference
    ANI <  95 and AF >= 0.60  -> putative novel species-level genome
    ANI <  95 and AF <  0.60  -> divergent from available references; species
                                 novelty likely but lower-confidence
    no close ANI hit          -> unresolved; evaluate with AAI/phylogeny

Writes in place; the original workbook is backed up to *.pre_s6.bak.xlsx first.
"""
import argparse
import json
import re
import shutil
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "MAG_target_gene_filled.xlsx"
SHEET = "Table S6_ANI"
RANKS = ["Domain", "Phylum", "Class", "Order", "Family", "Genus", "Species"]
RANK_PREFIX = dict(zip(RANKS, ["d__", "p__", "c__", "o__", "f__", "g__", "s__"]))
REF_DB = "GTDB R220 (GTDB-Tk v2.7.2, classify_wf)"

# S6 column letters (1-indexed openpyxl) -> meaning
COL = {
    "bin": 1, "iter": 2, "gtdb_full": 3,
    "Phylum": 4, "Class": 5, "Order": 6, "Family": 7, "Genus": 8, "Species": 9,
    "ref_name": 10, "ref_acc": 11, "ref_db": 12, "ani": 13, "af": 14,
    "qcov": 15, "rcov": 16, "call": 17, "novelty": 18, "table2": 19, "notes": 20,
}

# GTDB-Tk v2 summary column candidates (handle r2 closest_genome_* and legacy fastani_*)
ANI_COLS = ["closest_genome_ani", "fastani_ani"]
AF_COLS = ["closest_genome_af", "fastani_af"]
REFACC_COLS = ["closest_genome_reference", "fastani_reference"]
REFTAX_COLS = ["closest_genome_taxonomy", "fastani_taxonomy"]
METHOD_COLS = ["classification_method"]


def _norm(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none", "n/a", "") else s


def pick(row, names):
    for n in names:
        if n in row and _norm(row[n]):
            return _norm(row[n])
    return ""


def species_name_from_tax(tax_str):
    """Pull the s__ epithet out of a GTDB classification string."""
    for f in str(tax_str).split(";"):
        f = f.strip()
        if f.startswith("s__"):
            return f[3:].strip()
    return ""


RANK_WORD = {"d__": "domain", "p__": "phylum", "c__": "class", "o__": "order",
             "f__": "family", "g__": "genus", "s__": "species"}


def gtdb_species_and_depth(classif):
    """(species_epithet_or_'', deepest_defined_rank_word) from a GTDB string."""
    sp, deepest = "", "domain"
    for f in str(classif).split(";"):
        f = f.strip()
        for pfx, word in RANK_WORD.items():
            if f.startswith(pfx) and len(f) > len(pfx):
                deepest = word
                if pfx == "s__":
                    sp = f[len(pfx):].strip()
    return sp, deepest


def acc_from_ref_file(ref_file):
    """GCF_000007185.1_genomic.fna.gz -> GCF_000007185.1 (robust to dirs/exts)."""
    base = Path(str(ref_file)).name
    base = re.sub(r"\.(fna|fa|fasta)(\.gz)?$", "", base)
    base = re.sub(r"_genomic$", "", base)
    return base


def load_skani_best(path):
    """skani 'search'/'dist' output -> {mag: {ani, af_q, af_r, ref_acc}} (best ANI).

    Expects skani columns Ref_file, Query_file, ANI, Align_fraction_ref,
    Align_fraction_query (the default header). Query basename (minus extension)
    is the MAG name; reference accession is parsed from Ref_file."""
    sk = pd.read_csv(path, sep="\t")
    out = {}
    for _, r in sk.iterrows():
        mag = re.sub(r"\.(fa|fna|fasta)(\.gz)?$", "", Path(str(r["Query_file"])).name)
        ani = float(r["ANI"])
        cur = out.get(mag)
        if cur is None or ani > cur["ani"]:
            out[mag] = {
                "ani": ani,
                "af_q": float(r.get("Align_fraction_query", "nan")),
                "af_r": float(r.get("Align_fraction_ref", "nan")),
                "ref_acc": acc_from_ref_file(r["Ref_file"]),
            }
    return out


def novelty(ani, af, classif=""):
    """Return (species_level_call, novelty_interpretation).

    Implements the collaborator's 4 rules, plus an explicit ANI>=95/AF<0.60 edge
    arm, and enriches the no-ANI-hit ("unresolved") arm with GTDB's own
    species-novelty signal (whether GTDB assigned a species via RED/topology)."""
    if ani == "" or ani is None:
        sp, deepest = gtdb_species_and_depth(classif)
        if sp:
            return ("Unresolved by ANI (GTDB species assigned by RED/topology)",
                    "No close ANI hit; GTDB assigns a species via RED/topology; "
                    "confirm with AAI/phylogeny.")
        return ("Unresolved by ANI (GTDB species novel; s__ unassigned)",
                f"No close ANI hit and GTDB assigns no species (novel at species "
                f"level; GTDB-resolved to {deepest}); evaluate exact ANI/AF "
                f"(skani vs GTDB reps) or AAI/phylogeny.")
    ani = float(ani)
    has_af = af not in ("", None)
    afv = float(af) if has_af else None
    if ani >= 95 and has_af and afv >= 0.60:
        return ("Same species cluster",
                "Same species-level cluster as closest reference.")
    if ani < 95 and has_af and afv >= 0.60:
        return ("Putative novel species",
                "Putative novel species-level genome.")
    if ani < 95 and has_af and afv < 0.60:
        return ("Divergent / likely novel (lower confidence)",
                "Divergent from available references; species novelty likely "
                "but lower-confidence.")
    if ani >= 95 and has_af and afv < 0.60:  # edge: high ANI over short alignment
        return ("Same species (AF<0.60, lower confidence)",
                "ANI >=95% to closest reference but AF <0.60 (limited "
                "alignment); same-species assignment supported by ANI but "
                "lower-confidence.")
    # ANI present, AF entirely missing (not seen in this dataset)
    return ("Same species cluster (AF unavailable)" if ani >= 95
            else "Putative novel species (AF unavailable)",
            "ANI-only call; AF unavailable.")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--gtdbtk", default=None,
                    help="path to gtdbtk.bac120.summary.tsv (optional)")
    ap.add_argument("--skani", default=None,
                    help="path to skani search output (MAGs vs GTDB r220 reps); "
                         "fills ANI/AF/coverage for MAGs lacking a GTDB-Tk ANI")
    ap.add_argument("--skani-all", action="store_true",
                    help="use skani for ALL MAGs (overwrite GTDB-Tk ANI rows too)")
    ap.add_argument("--gtdb-tax", default=None,
                    help="GTDB accession->taxonomy tsv (to name skani closest refs)")
    ap.add_argument("--xlsx", default=str(XLSX))
    args = ap.parse_args()
    xlsx = Path(args.xlsx)

    # ---- local sources ----
    iter_map = json.loads((ROOT / "mag_iterativeID_old_to_new.json").read_text())
    qual = pd.read_csv(ROOT / "mag_quality_summary.tsv", sep="\t").set_index("MAG")

    # locate GTDB-Tk summary if not given
    gtdb = None
    cand = [args.gtdbtk] if args.gtdbtk else [
        ROOT / "gtdbtk.bac120.summary.tsv",
        ROOT / "meta" / "gtdbtk.bac120.summary.tsv",
    ]
    for c in cand:
        if c and Path(c).exists():
            gtdb = pd.read_csv(c, sep="\t").set_index("user_genome")
            print(f"GTDB-Tk summary loaded: {c}  ({len(gtdb)} rows, cols={list(gtdb.columns)[:6]}...)")
            break
    if gtdb is None:
        print("NOTE: GTDB-Tk summary NOT found -> ANI columns (J,K,M,N,Q,R) left blank.\n"
              "      Provide it via --gtdbtk and re-run to complete the table.")

    # optional skani (fills the no-ANI MAGs + the two coverage columns)
    skani = load_skani_best(args.skani) if args.skani else {}
    if skani:
        print(f"skani loaded: {len(skani)} MAGs with a nearest-reference ANI.")
    # optional accession -> GTDB species name lookup (for skani closest-ref names)
    acc2sp = {}
    if args.gtdb_tax and Path(args.gtdb_tax).exists():
        tx = pd.read_csv(args.gtdb_tax, sep="\t", header=None, names=["acc", "tax"])
        acc2sp = {a: species_name_from_tax(t) for a, t in zip(tx["acc"], tx["tax"])}
        # GTDB taxonomy files key on accessions like GB_GCA_... / RS_GCF_...; index both
        acc2sp.update({a.split("_", 1)[1]: s for a, s in list(acc2sp.items())
                       if a[:3] in ("GB_", "RS_")})
        print(f"GTDB taxonomy lookup: {len(acc2sp)} accessions.")

    # ---- same-GTDB-species co-member groups (for the redundancy note) ----
    sp_groups = {}
    for mag, r in qual.iterrows():
        sp = _norm(r.get("Species"))
        if sp:
            sp_groups.setdefault(sp, []).append(mag)
    sp_groups = {k: v for k, v in sp_groups.items() if len(v) > 1}

    # ---- write ----
    bak = xlsx.with_suffix(".pre_s6.bak.xlsx")
    if not bak.exists():
        shutil.copy2(xlsx, bak)
        print(f"Backup written: {bak.name}")
    wb = openpyxl.load_workbook(xlsx)
    ws = wb[SHEET]

    n_local = n_ani = n_unres = n_skani = 0
    for row in range(2, ws.max_row + 1):
        mag = _norm(ws.cell(row, COL["bin"]).value)
        if not mag:
            continue
        # --- local columns ---
        ws.cell(row, COL["iter"]).value = iter_map.get(mag, "")
        if mag in qual.index:
            q = qual.loc[mag]
            ranks = {rk: _norm(q.get(rk)) for rk in RANKS}
            for rk in ["Phylum", "Class", "Order", "Family", "Genus", "Species"]:
                ws.cell(row, COL[rk]).value = ranks[rk]
            full = ";".join(f"{RANK_PREFIX[rk]}{ranks[rk]}" for rk in RANKS)
            ws.cell(row, COL["gtdb_full"]).value = full
        ws.cell(row, COL["ref_db"]).value = REF_DB
        n_local += 1

        # same-species co-member note (always available locally)
        notes = []
        sp = _norm(q.get("Species")) if mag in qual.index else ""
        if sp in sp_groups:
            others = [m for m in sp_groups[sp] if m != mag]
            notes.append(
                f"Distinct genome within GTDB species s__{sp}; co-member(s): "
                f"{', '.join(others)}. Separate dRep 95% representatives "
                f"(same GTDB species != same dRep cluster); not a dereplication "
                f"error."
            )

        # --- ANI columns from GTDB-Tk summary ---
        had_gtdb_ani = False
        if gtdb is not None and mag in gtdb.index:
            gr = gtdb.loc[mag].to_dict()
            ani = pick(gr, ANI_COLS)
            af = pick(gr, AF_COLS)
            ref_acc = pick(gr, REFACC_COLS)
            ref_tax = pick(gr, REFTAX_COLS)
            method = pick(gr, METHOD_COLS)
            had_gtdb_ani = ani != ""
            ws.cell(row, COL["ref_acc"]).value = ref_acc
            ws.cell(row, COL["ref_name"]).value = species_name_from_tax(ref_tax) or ref_tax
            ws.cell(row, COL["ani"]).value = float(ani) if ani else None
            ws.cell(row, COL["af"]).value = float(af) if af else None
            call, interp = novelty(ani, af, pick(gr, ["classification"]))
            ws.cell(row, COL["call"]).value = call
            ws.cell(row, COL["novelty"]).value = interp
            if ani == "":
                n_unres += 1
            n_ani += 1
            if method:
                notes.append(f"GTDB-Tk classification_method: {method}.")
            note_val = pick(gr, ["note"])
            if note_val:
                notes.append(f"GTDB-Tk note: {note_val}.")

        # --- skani fallback: fill no-ANI MAGs (or all, with --skani-all) ---
        if skani and mag in skani and (args.skani_all or not had_gtdb_ani):
            s = skani[mag]
            af_q = s["af_q"] if pd.notna(s["af_q"]) else None     # percent
            af_r = s["af_r"] if pd.notna(s["af_r"]) else None     # percent
            af_frac = (af_q / 100.0) if af_q is not None else ""   # fraction for rule + col N
            classif = pick(gtdb.loc[mag].to_dict(), ["classification"]) if (
                gtdb is not None and mag in gtdb.index) else ""
            ws.cell(row, COL["ref_acc"]).value = s["ref_acc"]
            ws.cell(row, COL["ref_name"]).value = acc2sp.get(s["ref_acc"], "") or s["ref_acc"]
            ws.cell(row, COL["ani"]).value = round(s["ani"], 2)
            ws.cell(row, COL["af"]).value = round(af_frac, 3) if af_frac != "" else None
            ws.cell(row, COL["qcov"]).value = round(af_q, 1) if af_q is not None else None
            ws.cell(row, COL["rcov"]).value = round(af_r, 1) if af_r is not None else None
            call, interp = novelty(f"{s['ani']:.2f}", af_frac, classif)
            ws.cell(row, COL["call"]).value = call
            ws.cell(row, COL["novelty"]).value = interp
            if not had_gtdb_ani:
                n_unres -= 1                       # this MAG is no longer unresolved
                n_skani += 1
            notes.append("ANI/AF/coverage from skani vs GTDB r220 species reps "
                         "(nearest representative).")

        if notes:
            ws.cell(row, COL["notes"]).value = " ".join(notes)

    wb.save(xlsx)
    print(f"\nWrote {SHEET}: {n_local} MAGs local-filled; "
          f"{n_ani} GTDB-Tk-ANI-filled; {n_skani} resolved via skani; "
          f"{max(n_unres,0)} still unresolved/no-hit.")
    print(f"Same-GTDB-species multi-member groups: {len(sp_groups)} "
          f"(e.g. {list(sp_groups.items())[:2]})")
    print(f"Saved: {xlsx}")


if __name__ == "__main__":
    main()
