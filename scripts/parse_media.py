#!/usr/bin/env python
"""
Parse Recipe 2 ("Synthetic wastewater media, free of VFA, P, NO3-N & NO2-N")
from `Synthetic media and other reagents preparation_EM.xlsx`, sheet
'Media ingredients_conc.' (rows 31-35).

Outputs (mM as the concentration unit, per user choice):
  1. media_recipe2_concentrations.json
        - compounds_as_parsed : each chemical with formula, category, the
          source mg/L from the sheet, the dosing factor applied, the molar
          mass used, and the resulting mM in the final medium.
        - species_resolved    : salts split into their constituent ions with
          molar concentrations SUMMED across every salt that contributes the
          ion; non-salts (boric acid, EDTA, ATU, yeast extract) kept whole;
          water of crystallisation reported separately (it is not an ion).
  2. media_recipe2_modelseed_ids.json
        - species_to_modelseed : {species -> cpd##### or null}
        - match_details        : how each match was made (formula+charge / name)
        - parent_salts_to_modelseed : best-effort mapping of the intact salts.

Trace-element stocks are dosed at 0.3 mL per L of medium -> factor 0.3/1000.
"""
import json, csv, re, glob, os
from collections import defaultdict

XLSX = "Synthetic media and other reagents preparation_EM.xlsx"
SHEET = "Media ingredients_conc."
MSD_GLOB = "/Users/andrewfreiburger/Documents/Research/ModelSEEDDatabase/Biochemistry/compound_*.tsv"
TRACE_DOSING_ML_PER_L = 0.3          # trace-elements solution dose
TRACE_FACTOR = TRACE_DOSING_ML_PER_L / 1000.0

# ---- standard atomic weights (IUPAC abridged) ---------------------------------
AW = {
    'H':1.008,'B':10.811,'C':12.011,'N':14.007,'O':15.999,'Na':22.990,
    'Mg':24.305,'S':32.06,'Cl':35.45,'K':39.098,'Ca':40.078,'Mn':54.938,
    'Fe':55.845,'Co':58.933,'Ni':58.693,'Cu':63.546,'Zn':65.38,'Mo':95.95,
    'I':126.904,
}

# ---- formula parsing ----------------------------------------------------------
_token = re.compile(r'([A-Z][a-z]?)(\d*)')

def _counts_simple(frag):
    """element counts for a formula fragment with parentheses support."""
    counts = defaultdict(float)
    i, stack = 0, [defaultdict(float)]
    while i < len(frag):
        c = frag[i]
        if c == '(':
            stack.append(defaultdict(float)); i += 1
        elif c == ')':
            i += 1
            m = re.match(r'(\d+)', frag[i:]); mult = int(m.group(1)) if m else 1
            if m: i += m.end()
            top = stack.pop()
            for el, n in top.items(): stack[-1][el] += n * mult
        else:
            m = _token.match(frag, i)
            if not m or not m.group(1):
                i += 1; continue
            el = m.group(1); n = int(m.group(2)) if m.group(2) else 1
            stack[-1][el] += n; i = m.end()
    base = stack[0]
    for el, n in base.items(): counts[el] += n
    return counts

def parse_formula(formula):
    """Return (element_counts_including_hydrate, n_water_of_hydration).
    Handles 'MgSO4·7H2O' style hydrates (·, *, .)."""
    parts = re.split(r'[·*.]', formula)
    main = parts[0]
    counts = _counts_simple(main)
    n_water = 0.0
    for extra in parts[1:]:
        m = re.match(r'(\d*)\s*H2O$', extra.strip())
        if m:
            k = int(m.group(1)) if m.group(1) else 1
            n_water += k
            counts['H'] += 2*k; counts['O'] += k
        else:                       # generic hydrate/adduct fragment
            sub = _counts_simple(re.sub(r'^\d+', '', extra))
            mult = int(re.match(r'(\d*)', extra).group(1) or 1)
            for el, n in sub.items(): counts[el] += n*mult
    return counts, n_water

def mol_weight(formula):
    counts, _ = parse_formula(formula)
    return sum(AW[el]*n for el, n in counts.items())

def signature(formula):
    """anhydrous element-count signature for matching (ignores hydrate)."""
    counts, _ = parse_formula(formula)
    return tuple(sorted((el, int(round(n))) for el, n in counts.items() if n))

# ---- read Recipe 2 from the workbook ------------------------------------------
import openpyxl
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb[SHEET]
NAME_ROW, FORMULA_ROW, CONC_ROW = 33, 34, 35
# columns 2-7 = major components ; 8-17 = trace elements (per merged headers)
FORMULA_OVERRIDE = {                       # sheet 'formula' missing or not a real formula
    'allylthiourea (ATU)': 'C4H8N2S',
    'Ethylenediaminetetraacetic acid': 'C10H16N2O8',   # free acid (dissolved with NaOH)
    'Yeast extract': None,
}
compounds = {}
for col in range(2, 18):
    name = ws.cell(row=NAME_ROW, column=col).value
    if not name:
        continue
    name = str(name).strip()
    raw_formula = ws.cell(row=FORMULA_ROW, column=col).value
    conc = ws.cell(row=CONC_ROW, column=col).value
    if conc is None:
        continue
    formula = FORMULA_OVERRIDE.get(name, raw_formula)
    if formula is not None:
        formula = str(formula).strip()
    category = 'major' if col <= 7 else 'trace'
    factor = 1.0 if category == 'major' else TRACE_FACTOR
    final_mg_L = float(conc) * factor
    mw = mol_weight(formula) if formula else None
    mM = (final_mg_L / mw) if mw else None
    compounds[name] = {
        'formula': formula,
        'category': category,
        'source_conc_mg_L': float(conc),
        'dosing_mL_per_L': None if category == 'major' else TRACE_DOSING_ML_PER_L,
        'final_conc_mg_L': round(final_mg_L, 6),
        'MW_g_per_mol': round(mw, 4) if mw else None,
        'mM': round(mM, 8) if mM is not None else None,
    }

# ---- dissociation rules (keyed by anhydrous formula) --------------------------
# ion key -> (formula, charge) for the constituent ions
ION_DEF = {
    'NH4+':   ('NH4', +1), 'Cl-':    ('Cl', -1), 'Mg2+': ('Mg', +2),
    'SO4_2-': ('SO4', -2), 'Ca2+':   ('Ca', +2), 'Na+':  ('Na', +1),
    'HCO3-':  ('HCO3', -1), 'Fe3+':  ('Fe', +3), 'Cu2+': ('Cu', +2),
    'K+':     ('K', +1),   'I-':     ('I', -1),  'Mn2+': ('Mn', +2),
    'MoO4_2-':('MoO4', -2), 'Zn2+':  ('Zn', +2), 'Co2+': ('Co', +2),
}
DISSOCIATION = {        # anhydrous formula -> [(ion_key, stoichiometry), ...]
    'NH4Cl':   [('NH4+',1), ('Cl-',1)],
    'MgSO4':   [('Mg2+',1), ('SO4_2-',1)],
    'CaCl2':   [('Ca2+',1), ('Cl-',2)],
    'NaHCO3':  [('Na+',1),  ('HCO3-',1)],
    'FeCl3':   [('Fe3+',1), ('Cl-',3)],
    'CuSO4':   [('Cu2+',1), ('SO4_2-',1)],
    'KI':      [('K+',1),   ('I-',1)],
    'MnCl2':   [('Mn2+',1), ('Cl-',2)],
    'Na2MoO4': [('Na+',2),  ('MoO4_2-',1)],
    'ZnSO4':   [('Zn2+',1), ('SO4_2-',1)],
    'CoCl2':   [('Co2+',1), ('Cl-',2)],
}
KEEP_WHOLE = {'H3BO3', 'C10H16N2O8', 'C4H8N2S', None}   # boric acid, EDTA, ATU, yeast

ion_mM = defaultdict(float)
ion_sources = defaultdict(list)
molecules_whole = {}
water_mM = 0.0
for name, c in compounds.items():
    formula = c['formula']
    anhydrous = re.split(r'[·*.]', formula)[0] if formula else None
    if anhydrous in DISSOCIATION:
        _, n_water = parse_formula(formula)
        for ion_key, stoich in DISSOCIATION[anhydrous]:
            ion_mM[ion_key] += stoich * c['mM']
            ion_sources[ion_key].append(name)
        water_mM += n_water * c['mM']
    else:                                   # kept whole (organics / weak acids)
        molecules_whole[name] = c['mM']

# ---- assemble concentrations JSON ---------------------------------------------
ions_out = {}
for ion_key in sorted(ion_mM, key=lambda k: -ion_mM[k]):
    f, ch = ION_DEF[ion_key]
    ions_out[ion_key] = {
        'formula': f, 'charge': ch,
        'mM': round(ion_mM[ion_key], 8),
        'from_salts': sorted(set(ion_sources[ion_key])),
    }
molecules_out = {n: (round(v, 8) if v is not None else None)
                 for n, v in molecules_whole.items()}

concentrations = {
    'meta': {
        'recipe': 'Synthetic wastewater media (free of VFA, P, NO3--N and NO2--N)',
        'source_file': XLSX,
        'source_sheet': SHEET,
        'source_rows': '31-35 (Recipe 2)',
        'concentration_unit': 'mM',
        'trace_element_dosing_mL_per_L': TRACE_DOSING_ML_PER_L,
        'notes': [
            'Trace-element concentrations are the in-medium values after the '
            '0.3 mL/L dose (stock mg/L x 0.3/1000).',
            'Salts dissociated into constituent ions; molar contributions summed '
            'across all salts sharing an ion.',
            'Boric acid, EDTA (free acid C10H16N2O8), allylthiourea (ATU) and '
            'yeast extract are kept whole (not ionised). Yeast extract has no '
            'defined formula so no molar concentration is computed.',
            'Water of crystallisation is reported separately; it merges into the '
            'bulk solvent and is not an ionic species.',
        ],
    },
    'compounds_as_parsed': compounds,
    'species_resolved': {
        'ions': ions_out,
        'molecules_kept_whole': molecules_out,
        'water_of_hydration_released_mM': round(water_mM, 6),
    },
}
with open('media_recipe2_concentrations.json', 'w') as fh:
    json.dump(concentrations, fh, indent=2, ensure_ascii=False)

# ---- load ModelSEED compounds (clean rows only) -------------------------------
by_id = {}
sig_charge_index = defaultdict(list)     # (signature, charge) -> [id...]
name_index = defaultdict(list)           # lowercased name/alias -> [id...]
for path in glob.glob(MSD_GLOB):
    with open(path, newline='') as fh:
        r = csv.reader(fh, delimiter='\t')
        header = next(r, None)
        for row in r:
            if len(row) < 10 or not re.match(r'^cpd\d{5,}$', row[0]):
                continue
            cid, name, formula = row[0], row[2], row[3]
            charge_s, is_obs = row[7], row[9]
            if name.startswith(('MetaCyc:', 'KEGG:', 'ChEBI:')) or ';' in name:
                continue                 # alias / shifted row
            try:
                charge = int(charge_s)
            except (ValueError, TypeError):
                continue
            if abs(charge) > 1000:       # sentinel 10000000 = unknown
                charge = None
            if cid in by_id:
                continue
            aliases = row[18] if len(row) > 18 else ''
            by_id[cid] = {'name': name, 'formula': formula, 'mass': row[4],
                          'charge': charge, 'is_obsolete': is_obs}
            name_index[name.lower().strip()].append(cid)
            for al in re.split(r'[;|]', aliases):
                al = al.strip().lower()
                if al and ':' not in al:
                    name_index[al].append(cid)
            if formula and formula != 'null' and charge is not None:
                try:
                    sig_charge_index[(signature(formula), charge)].append(cid)
                except Exception:
                    pass

def pick(ids):
    if not ids: return None
    ids = sorted(set(ids), key=lambda c: (by_id[c]['is_obsolete'] not in ('0', 0),
                                          int(re.sub(r'\D', '', c))))
    return ids[0]

def match_species(formula, charge, name_hints):
    # 1) exact element-signature + charge
    if formula:
        try:
            cands = sig_charge_index.get((signature(formula), charge), [])
            cid = pick(cands)
            if cid:
                return cid, 'formula+charge'
        except Exception:
            pass
    # 2) name / alias
    for h in name_hints:
        cid = pick(name_index.get(h.lower().strip(), []))
        if cid:
            return cid, 'name'
    return None, None

# ---- map ions + molecules -----------------------------------------------------
NAME_HINTS = {                         # extra name hints for hard cases
    'NH4+': ['ammonium', 'nh4', 'ammonia'],
    'HCO3-': ['bicarbonate', 'hydrogen carbonate'],
    'MoO4_2-': ['molybdate'],
    'I-': ['iodide'],
    'SO4_2-': ['sulfate', 'sulphate'],
}
MOLECULE_DEF = {                       # whole molecules: (formula, charge, hints)
    'allylthiourea (ATU)': ('C4H8N2S', 0, ['allylthiourea']),
    'Boric acid':          ('H3BO3', 0, ['boric acid', 'borate']),
    'Ethylenediaminetetraacetic acid': ('C10H16N2O8', 0,
                                         ['edta', 'ethylenediaminetetraacetic acid']),
    'Yeast extract':       (None, None, ['yeast extract']),
}
species_to_id, details = {}, {}
for ion_key, (f, ch) in ION_DEF.items():
    if ion_key not in ions_out:        # only those actually present
        continue
    cid, how = match_species(f, ch, NAME_HINTS.get(ion_key, []))
    species_to_id[ion_key] = cid
    details[ion_key] = ({'id': cid, 'ms_name': by_id[cid]['name'],
                         'ms_formula': by_id[cid]['formula'],
                         'ms_charge': by_id[cid]['charge'], 'matched_by': how}
                        if cid else {'id': None, 'matched_by': None})
# molecules that were kept whole
present_mol_names = set(molecules_whole)
for disp, (f, ch, hints) in MOLECULE_DEF.items():
    # align display name to whatever appears in the sheet's molecule set
    key = next((n for n in present_mol_names
                if disp.split(' (')[0].lower() in n.lower()
                or n.lower() in disp.lower()), disp)
    cid, how = match_species(f, ch, hints)
    species_to_id[key] = cid
    details[key] = ({'id': cid, 'ms_name': by_id[cid]['name'],
                     'ms_formula': by_id[cid]['formula'],
                     'ms_charge': by_id[cid]['charge'], 'matched_by': how}
                    if cid else {'id': None, 'matched_by': None})
# water of hydration
cid, how = match_species('H2O', 0, ['h2o', 'water'])
species_to_id['H2O (water of hydration)'] = cid
details['H2O (water of hydration)'] = {'id': cid, 'ms_name': by_id[cid]['name'],
    'ms_formula': by_id[cid]['formula'], 'ms_charge': by_id[cid]['charge'],
    'matched_by': how} if cid else {'id': None}

# ---- bonus: map intact parent salts by name -----------------------------------
parent_salts = {}
for name, c in compounds.items():
    cid = pick(name_index.get(name.lower().strip(), []))
    if not cid and c['formula']:       # try formula+charge 0 (neutral salt)
        cid, _ = match_species(c['formula'], 0, [name])
    parent_salts[name] = cid

mapping = {
    'meta': {
        'modelseed_database': os.path.dirname(MSD_GLOB),
        'match_strategy': 'element-signature + charge, then name/alias',
        'notes': [
            'IDs are ModelSEED compound IDs (cpd#####).',
            'Ion IDs are the charged metabolite forms; parent_salts are the '
            'neutral reagent compounds (best effort, may be null).',
            'EDTA maps to the cpd00240 EDTA(4-) metabolite (ModelSEED has no '
            'free-acid form); matched by name.',
            'Yeast extract has no single ModelSEED compound (null).',
        ],
    },
    'species_to_modelseed': species_to_id,
    'match_details': details,
    'parent_salts_to_modelseed': parent_salts,
}
with open('media_recipe2_modelseed_ids.json', 'w') as fh:
    json.dump(mapping, fh, indent=2, ensure_ascii=False)

# ---- report -------------------------------------------------------------------
print(f"Loaded {len(by_id)} clean ModelSEED compounds.\n")
print("=== COMPOUNDS AS PARSED (final mM) ===")
for n, c in compounds.items():
    print(f"  {n:42s} {str(c['formula']):14s} {c['category']:5s} "
          f"{c['final_conc_mg_L']:>10} mg/L  MW={c['MW_g_per_mol']}  mM={c['mM']}")
print("\n=== SUMMED IONS (mM) ===")
for k, v in ions_out.items():
    print(f"  {k:9s} {v['mM']:<14}  from {v['from_salts']}")
print("\n=== MOLECULES KEPT WHOLE (mM) ===")
for k, v in molecules_out.items():
    print(f"  {k:42s} {v}")
print(f"\n  water of hydration released: {water_mM:.4f} mM")
print("\n=== ModelSEED MAPPING ===")
for k, v in species_to_id.items():
    d = details.get(k, {})
    print(f"  {k:34s} -> {str(v):10s} "
          f"({d.get('ms_name','-')} | {d.get('ms_formula','-')} | "
          f"chg={d.get('ms_charge','-')} | by {d.get('matched_by','-')})")
print("\n  parent salts:")
for k, v in parent_salts.items():
    print(f"    {k:42s} -> {v}")
print("\nWrote media_recipe2_concentrations.json and media_recipe2_modelseed_ids.json")
